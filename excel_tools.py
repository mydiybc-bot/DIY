from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook

from data_store import TrainingContentStore


RULE_HEADERS = ["key", "value"]
QUESTION_HEADERS = [
    "section_id",
    "section_title",
    "question_number",
    "prompt",
    "answer",
    "point_label",
    "keywords",
]
REPORT_HEADERS = [
    "created_at",
    "role",
    "employee_id",
    "employee_name",
    "section_id",
    "section_title",
    "question_count",
    "scores",
    "average_score",
]


def export_content_to_excel(content: dict) -> bytes:
    workbook = Workbook()

    rules_sheet = workbook.active
    rules_sheet.title = "rules"
    rules_sheet.append(RULE_HEADERS)
    for key, value in content["rules"].items():
        rules_sheet.append([key, value])

    questions_sheet = workbook.create_sheet("questions")
    questions_sheet.append(QUESTION_HEADERS)
    for section in content["sections"]:
        for question in section["questions"]:
            for point in question["required_points"]:
                questions_sheet.append(
                    [
                        section["id"],
                        section["title"],
                        question["number"],
                        question["prompt"],
                        question["answer"],
                        point["label"],
                        ", ".join(point["keywords"]),
                    ]
                )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def import_content_from_excel(raw_bytes: bytes, store: TrainingContentStore) -> dict:
    workbook = load_workbook(BytesIO(raw_bytes))
    if "rules" not in workbook.sheetnames or "questions" not in workbook.sheetnames:
        raise ValueError("Excel 檔需要包含 rules 與 questions 兩個工作表")

    rules_sheet = workbook["rules"]
    question_sheet = workbook["questions"]

    rules = {}
    for row in rules_sheet.iter_rows(min_row=2, values_only=True):
        key, value = row[:2]
        if key:
            rules[str(key).strip()] = "" if value is None else str(value)

    sections_map: dict[str, dict] = {}
    for row_index, row in enumerate(question_sheet.iter_rows(min_row=2, values_only=True), start=2):
        (
            section_id,
            section_title,
            question_number,
            prompt,
            answer,
            point_label,
            keywords,
        ) = row[:7]

        if not any(row[:7]):
            continue

        if not section_title or not prompt or not answer or not point_label:
            raise ValueError(f"questions 工作表第 {row_index} 列有空白必填欄位")

        normalized_section_id = str(section_id).strip() if section_id else f"section_{len(sections_map) + 1}"
        if normalized_section_id not in sections_map:
            sections_map[normalized_section_id] = {
                "id": normalized_section_id,
                "title": str(section_title).strip(),
                "questions": {},
            }

        question_key = int(question_number) if question_number else len(sections_map[normalized_section_id]["questions"]) + 1
        section = sections_map[normalized_section_id]
        if question_key not in section["questions"]:
            section["questions"][question_key] = {
                "prompt": str(prompt).strip(),
                "answer": str(answer).strip(),
                "required_points": [],
            }

        section["questions"][question_key]["required_points"].append(
            {
                "label": str(point_label).strip(),
                "keywords": [item.strip() for item in str(keywords or "").split(",") if item.strip()],
            }
        )

    sections = []
    for section in sections_map.values():
        ordered_questions = []
        for _, question in sorted(section["questions"].items(), key=lambda item: item[0]):
            ordered_questions.append(question)
        sections.append({"id": section["id"], "title": section["title"], "questions": ordered_questions})

    return store.validate({"rules": rules, "sections": sections})


def export_reports_to_excel(reports: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "reports"
    sheet.append(REPORT_HEADERS)
    for report in reports:
        sheet.append(
            [
                report.get("created_at", ""),
                report.get("role", ""),
                report.get("employee_id", ""),
                report.get("employee_name", ""),
                report.get("section_id", ""),
                report.get("section_title", ""),
                report.get("question_count", 0),
                "；".join(report.get("scores", [])),
                report.get("average_score", 0),
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
