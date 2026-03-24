from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime
import json
from pathlib import Path
from threading import Lock

from openpyxl import load_workbook


SOURCE_PATH = Path("/Users/diybc/Desktop/Claude/自己人/自己人消費記錄分析0323.xlsx")
SNAPSHOT_PATH = Path(__file__).parent / "data" / "self_dashboard_snapshot.json"
CLOSED_STORE = "台南新光西門店"
ACTIVE_STORES = [
    "台北南京店",
    "台南Focus店",
    "台中草悟道店",
    "台北士林店",
    "新北板橋店",
    "台北遠百信義A13店",
    "台中精明店",
    "桃園中壢店",
    "新竹文化店",
    "高雄SKM Park店",
    "新北新店店",
    "桃園藝文店",
]
ESTIMATED_DESSERT_PRICE = 550

_CACHE_LOCK = Lock()
_CACHE: dict[str, object] = {"mtime": None, "dataset": None}


def load_self_dashboard(filters: dict[str, str] | None = None) -> dict:
    dataset = _load_dataset()
    return _build_payload(dataset, filters or {})


def _load_dataset() -> dict:
    if SNAPSHOT_PATH.exists():
        snapshot_mtime = SNAPSHOT_PATH.stat().st_mtime
        with _CACHE_LOCK:
            if _CACHE["mtime"] == snapshot_mtime and _CACHE["dataset"] is not None:
                return _CACHE["dataset"]  # type: ignore[return-value]

            dataset = json.loads(SNAPSHOT_PATH.read_text(encoding="utf-8"))
            _CACHE["mtime"] = snapshot_mtime
            _CACHE["dataset"] = dataset
            return dataset

    source_mtime = SOURCE_PATH.stat().st_mtime
    with _CACHE_LOCK:
        if _CACHE["mtime"] == source_mtime and _CACHE["dataset"] is not None:
            return _CACHE["dataset"]  # type: ignore[return-value]

        dataset = _parse_source_workbook()
        _CACHE["mtime"] = source_mtime
        _CACHE["dataset"] = dataset
        return dataset


def export_self_dashboard_snapshot(output_path: Path | None = None) -> Path:
    target = output_path or SNAPSHOT_PATH
    target.parent.mkdir(parents=True, exist_ok=True)
    dataset = _parse_source_workbook()
    target.write_text(json.dumps(dataset, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    with _CACHE_LOCK:
        _CACHE["mtime"] = target.stat().st_mtime
        _CACHE["dataset"] = dataset
    return target


def _parse_source_workbook() -> dict:
    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    raw_sheet = workbook["自己人原始資料"]
    revenue_sheet = workbook["各店月營收"]

    join_daily: dict[tuple[str, str], dict[str, int]] = defaultdict(
        lambda: {"join_self": 0, "upgrade_self": 0, "group_join": 0}
    )
    consume_daily: dict[tuple[str, str], dict[str, int]] = defaultdict(
        lambda: {"desserts": 0, "consume_rows": 0, "coupon_use": 0, "coupon_100": 0}
    )
    pair_daily: Counter[tuple[str, str, str]] = Counter()
    dessert_daily: Counter[tuple[str, str, str]] = Counter()
    coupon_daily: Counter[tuple[str, str, str]] = Counter()
    revenue_monthly: list[tuple[str, int, float]] = []

    date_min: str | None = None
    date_max: str | None = None
    revenue_latest_month = 0

    rows = raw_sheet.iter_rows(values_only=True)
    next(rows)
    for row in rows:
        if not any(value is not None for value in row):
            continue

        event_date = _row_date(row, 0)
        if not event_date:
            continue
        day_key = event_date.isoformat()
        date_min = day_key if date_min is None or day_key < date_min else date_min
        date_max = day_key if date_max is None or day_key > date_max else date_max

        join_store = _clean_text(_row_value(row, 1))
        description = _clean_text(_row_value(row, 3))
        consume_store = _clean_text(_row_value(row, 4))
        dessert = _clean_text(_row_value(row, 5))

        if join_store in ACTIVE_STORES:
            if description == "加入自己人":
                join_daily[(day_key, join_store)]["join_self"] += 1
            elif description == "升等 自己人":
                join_daily[(day_key, join_store)]["upgrade_self"] += 1
            elif description == "加入群友":
                join_daily[(day_key, join_store)]["group_join"] += 1

        if consume_store in ACTIVE_STORES:
            consume_daily[(day_key, consume_store)]["consume_rows"] += 1

            if dessert:
                consume_daily[(day_key, consume_store)]["desserts"] += 1
                dessert_daily[(day_key, consume_store, dessert)] += 1
                if join_store in ACTIVE_STORES:
                    pair_daily[(day_key, join_store, consume_store)] += 1

            if description.startswith("使用優惠券："):
                coupon_name = description.replace("使用優惠券：", "", 1).strip()
                consume_daily[(day_key, consume_store)]["coupon_use"] += 1
                if coupon_name == "100元甜點券":
                    consume_daily[(day_key, consume_store)]["coupon_100"] += 1
                coupon_daily[(day_key, consume_store, coupon_name)] += 1

    for row in revenue_sheet.iter_rows(min_row=2, values_only=True):
        store = _clean_text(_row_value(row, 0))
        ym_value = _to_int(_row_value(row, 4))
        revenue_value = _to_float(_row_value(row, 3))
        if store not in ACTIVE_STORES or ym_value is None or revenue_value is None:
            continue
        revenue_monthly.append((store, ym_value, revenue_value))
        revenue_latest_month = max(revenue_latest_month, ym_value)

    workbook.close()

    return {
        "source_path": str(SOURCE_PATH),
        "source_updated_at": datetime.fromtimestamp(SOURCE_PATH.stat().st_mtime).isoformat(timespec="seconds"),
        "date_min": date_min,
        "date_max": date_max,
        "default_start": _month_start(date_max) if date_max else None,
        "default_end": date_max,
        "revenue_latest_month": revenue_latest_month,
        "active_stores": ACTIVE_STORES,
        "closed_store": CLOSED_STORE,
        "join_daily": [
            {
                "date": day_key,
                "store": store,
                **metrics,
            }
            for (day_key, store), metrics in sorted(join_daily.items())
        ],
        "consume_daily": [
            {
                "date": day_key,
                "store": store,
                **metrics,
            }
            for (day_key, store), metrics in sorted(consume_daily.items())
        ],
        "pair_daily": [
            {
                "date": day_key,
                "join_store": join_store,
                "consume_store": consume_store,
                "desserts": count,
            }
            for (day_key, join_store, consume_store), count in sorted(pair_daily.items())
        ],
        "dessert_daily": [
            {
                "date": day_key,
                "store": store,
                "dessert": dessert,
                "count": count,
            }
            for (day_key, store, dessert), count in sorted(dessert_daily.items())
        ],
        "coupon_daily": [
            {
                "date": day_key,
                "store": store,
                "coupon": coupon,
                "count": count,
            }
            for (day_key, store, coupon), count in sorted(coupon_daily.items())
        ],
        "revenue_monthly": [
            {
                "store": store,
                "ym": ym_value,
                "revenue": revenue_value,
            }
            for store, ym_value, revenue_value in sorted(revenue_monthly)
        ],
    }


def _build_payload(dataset: dict, filters: dict[str, str]) -> dict:
    start_date = _coerce_date(filters.get("start_date")) or _coerce_date(dataset["default_start"])
    end_date = _coerce_date(filters.get("end_date")) or _coerce_date(dataset["default_end"])
    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    store_filter = filters.get("store", "全部") or "全部"
    if store_filter not in ACTIVE_STORES:
        store_filter = "全部"

    coupon_filter = filters.get("coupon", "全部") or "全部"
    known_coupons = sorted({item["coupon"] for item in dataset["coupon_daily"]})
    if coupon_filter != "全部" and coupon_filter not in known_coupons:
        coupon_filter = "全部"

    join_rows = [
        row for row in dataset["join_daily"]
        if _match_date(row["date"], start_date, end_date) and (store_filter == "全部" or row["store"] == store_filter)
    ]
    consume_rows = [
        row for row in dataset["consume_daily"]
        if _match_date(row["date"], start_date, end_date) and (store_filter == "全部" or row["store"] == store_filter)
    ]
    pair_rows = [
        row for row in dataset["pair_daily"]
        if _match_date(row["date"], start_date, end_date) and (store_filter == "全部" or row["join_store"] == store_filter)
    ]
    dessert_rows = [
        row for row in dataset["dessert_daily"]
        if _match_date(row["date"], start_date, end_date) and (store_filter == "全部" or row["store"] == store_filter)
    ]
    coupon_rows = [
        row
        for row in dataset["coupon_daily"]
        if _match_date(row["date"], start_date, end_date)
        and (store_filter == "全部" or row["store"] == store_filter)
        and (coupon_filter == "全部" or row["coupon"] == coupon_filter)
    ]
    coupon_rows_all = [
        row
        for row in dataset["coupon_daily"]
        if _match_date(row["date"], start_date, end_date)
        and (store_filter == "全部" or row["store"] == store_filter)
    ]

    join_totals = _totals_by_store(join_rows, "store", ["join_self", "upgrade_self", "group_join"])
    consume_totals = _totals_by_store(consume_rows, "store", ["desserts", "consume_rows", "coupon_use", "coupon_100"])
    revenue_totals = _revenue_totals(dataset["revenue_monthly"], start_date, end_date, store_filter)

    store_overview = []
    for store in ACTIVE_STORES:
        join_self = join_totals.get(store, {}).get("join_self", 0)
        upgrade_self = join_totals.get(store, {}).get("upgrade_self", 0)
        desserts = consume_totals.get(store, {}).get("desserts", 0)
        coupon_use = consume_totals.get(store, {}).get("coupon_use", 0)
        coupon_100 = consume_totals.get(store, {}).get("coupon_100", 0)
        revenue = revenue_totals.get(store, 0)
        estimated = desserts * ESTIMATED_DESSERT_PRICE
        store_overview.append(
            {
                "store": store,
                "join_self": join_self,
                "upgrade_self": upgrade_self,
                "join_total": join_self + upgrade_self,
                "desserts": desserts,
                "estimated_revenue": estimated,
                "revenue": revenue,
                "revenue_share": round(estimated / revenue, 4) if revenue else None,
                "coupon_use": coupon_use,
                "coupon_100": coupon_100,
            }
        )

    store_overview.sort(key=lambda item: (item["join_total"], item["desserts"]), reverse=True)

    monthly_trends = _build_monthly_trends(join_rows, consume_rows, coupon_rows_all)
    pair_summary = _build_pair_summary(pair_rows)
    dessert_rankings = _build_dessert_rankings(dessert_rows)
    coupon_summary = _build_coupon_summary(coupon_rows_all, coupon_filter)
    join_ranking = [
        {
            "store": item["store"],
            "join_self": item["join_self"],
            "upgrade_self": item["upgrade_self"],
            "join_total": item["join_total"],
        }
        for item in sorted(store_overview, key=lambda item: item["join_total"], reverse=True)
    ]

    join_self_total = sum(item["join_self"] for item in join_rows)
    upgrade_total = sum(item["upgrade_self"] for item in join_rows)
    group_join_total = sum(item["group_join"] for item in join_rows)
    dessert_total = sum(item["desserts"] for item in consume_rows)
    coupon_total = sum(item["coupon_use"] for item in consume_rows)
    coupon_100_total = sum(item["coupon_100"] for item in consume_rows)
    estimated_revenue = dessert_total * ESTIMATED_DESSERT_PRICE
    comparable_revenue = sum(revenue_totals.values())
    revenue_share = round(estimated_revenue / comparable_revenue, 4) if comparable_revenue else None

    return {
        "ok": True,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_path": dataset["source_path"],
        "source_updated_at": dataset["source_updated_at"],
        "filters": {
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "store": store_filter,
            "coupon": coupon_filter,
        },
        "meta": {
            "date_min": dataset["date_min"],
            "date_max": dataset["date_max"],
            "default_start": dataset["default_start"],
            "default_end": dataset["default_end"],
            "active_stores": ACTIVE_STORES,
            "closed_store": CLOSED_STORE,
            "estimated_dessert_price": ESTIMATED_DESSERT_PRICE,
            "revenue_latest_month": dataset["revenue_latest_month"],
            "available_coupons": ["全部", *known_coupons],
        },
        "summary": {
            "join_self_total": join_self_total,
            "upgrade_total": upgrade_total,
            "group_join_total": group_join_total,
            "dessert_total": dessert_total,
            "coupon_total": coupon_total,
            "coupon_100_total": coupon_100_total,
            "estimated_revenue": estimated_revenue,
            "comparable_revenue": comparable_revenue,
            "revenue_share": revenue_share,
            "same_store_desserts": pair_summary["same_store_desserts"],
            "cross_store_desserts": pair_summary["cross_store_desserts"],
            "cross_store_rate": pair_summary["cross_store_rate"],
        },
        "monthly_trends": monthly_trends,
        "store_overview": store_overview,
        "join_ranking": join_ranking,
        "pair_analysis": pair_summary,
        "dessert_rankings": dessert_rankings,
        "coupon_analysis": coupon_summary,
    }


def _build_monthly_trends(join_rows: list[dict], consume_rows: list[dict], coupon_rows: list[dict]) -> list[dict]:
    trend_map: dict[str, dict[str, int]] = defaultdict(lambda: {"join_self": 0, "upgrade_self": 0, "desserts": 0, "coupon_use": 0})

    for row in join_rows:
        month = row["date"][:7]
        trend_map[month]["join_self"] += row["join_self"]
        trend_map[month]["upgrade_self"] += row["upgrade_self"]
    for row in consume_rows:
        month = row["date"][:7]
        trend_map[month]["desserts"] += row["desserts"]
    for row in coupon_rows:
        month = row["date"][:7]
        trend_map[month]["coupon_use"] += row["count"]

    return [
        {
            "month": month,
            **trend_map[month],
        }
        for month in sorted(trend_map)
    ]


def _build_pair_summary(pair_rows: list[dict]) -> dict:
    top_pairs_counter: Counter[tuple[str, str]] = Counter()
    matrix = []
    matrix_counter: Counter[tuple[str, str]] = Counter()
    same_store = 0
    cross_store = 0

    for row in pair_rows:
        key = (row["join_store"], row["consume_store"])
        top_pairs_counter[key] += row["desserts"]
        matrix_counter[key] += row["desserts"]
        if row["join_store"] == row["consume_store"]:
            same_store += row["desserts"]
        else:
            cross_store += row["desserts"]

    top_pairs = [
        {
            "join_store": join_store,
            "consume_store": consume_store,
            "desserts": count,
        }
        for (join_store, consume_store), count in top_pairs_counter.most_common(12)
    ]

    for join_store in ACTIVE_STORES:
        row_values = {"join_store": join_store, "totals": []}
        for consume_store in ACTIVE_STORES:
            row_values["totals"].append(
                {
                    "consume_store": consume_store,
                    "desserts": matrix_counter[(join_store, consume_store)],
                }
            )
        matrix.append(row_values)

    total_desserts = same_store + cross_store
    return {
        "same_store_desserts": same_store,
        "cross_store_desserts": cross_store,
        "cross_store_rate": round(cross_store / total_desserts, 4) if total_desserts else None,
        "top_pairs": top_pairs,
        "matrix": matrix,
    }


def _build_dessert_rankings(dessert_rows: list[dict]) -> list[dict]:
    store_counters: dict[str, Counter[str]] = defaultdict(Counter)
    for row in dessert_rows:
        store_counters[row["store"]][row["dessert"]] += row["count"]

    rankings = []
    for store in ACTIVE_STORES:
        top_items = [
            {"dessert": dessert, "count": count}
            for dessert, count in store_counters[store].most_common(5)
        ]
        rankings.append({"store": store, "items": top_items})
    return rankings


def _build_coupon_summary(coupon_rows: list[dict], selected_coupon: str) -> dict:
    coupon_counter: Counter[str] = Counter()
    store_counter: Counter[str] = Counter()

    for row in coupon_rows:
        coupon_counter[row["coupon"]] += row["count"]
        store_counter[row["store"]] += row["count"]

    coupon_ranking = [
        {"coupon": coupon, "count": count}
        for coupon, count in coupon_counter.most_common(12)
    ]
    store_ranking = [
        {"store": store, "count": store_counter.get(store, 0)}
        for store in ACTIVE_STORES
    ]
    store_ranking.sort(key=lambda item: item["count"], reverse=True)

    return {
        "selected_coupon": selected_coupon,
        "coupon_ranking": coupon_ranking,
        "store_ranking": store_ranking,
    }


def _totals_by_store(rows: list[dict], store_key: str, fields: list[str]) -> dict[str, dict[str, int]]:
    totals: dict[str, dict[str, int]] = defaultdict(lambda: {field: 0 for field in fields})
    for row in rows:
        bucket = totals[row[store_key]]
        for field in fields:
            bucket[field] += int(row[field])
    return totals


def _revenue_totals(revenue_rows: list[dict], start_date: date | None, end_date: date | None, store_filter: str) -> dict[str, int]:
    start_ym = start_date.year * 100 + start_date.month if start_date else None
    end_ym = end_date.year * 100 + end_date.month if end_date else None
    totals: Counter[str] = Counter()

    for row in revenue_rows:
        if store_filter != "全部" and row["store"] != store_filter:
            continue
        ym_value = int(row["ym"])
        if start_ym and ym_value < start_ym:
            continue
        if end_ym and ym_value > end_ym:
            continue
        totals[row["store"]] += int(round(row["revenue"]))
    return dict(totals)


def _match_date(value: str, start_date: date | None, end_date: date | None) -> bool:
    current = _coerce_date(value)
    if current is None:
        return False
    if start_date and current < start_date:
        return False
    if end_date and current > end_date:
        return False
    return True


def _month_start(day_text: str | None) -> str | None:
    day = _coerce_date(day_text)
    if not day:
        return None
    return day.replace(day=1).isoformat()


def _coerce_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


def _clean_text(value) -> str:
    return "" if value is None else str(value).strip()


def _row_value(row: tuple, index: int):
    return row[index] if index < len(row) else None


def _row_date(row: tuple, index: int) -> date | None:
    value = _row_value(row, index)
    if isinstance(value, datetime):
        return value.date()
    return None


def _to_int(value) -> int | None:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _to_float(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
