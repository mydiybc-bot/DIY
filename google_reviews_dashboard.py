from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
import os
from pathlib import Path
import re
from statistics import mean

from openpyxl import load_workbook


DEFAULT_SOURCE_DIR = Path(
    os.environ.get(
        "GOOGLE_REVIEWS_SOURCE_DIR",
        "/Users/diybc/Desktop/Claude/Google 評論/原始 Google 評論整理資料",
    )
)
HEADER_KEYS = {"title", "stars", "text", "publishedatdate", "name"}
TEXT_HEADER_ORDER = ("title", "stars", "text", "publishedAtDate", "name")
DEFAULT_HEADERLESS_ORDER = ("name", "publishedAtDate", "stars", "text", "title")
POSITIVE_THEME_KEYWORDS = {
    "服務協助": ["幫忙", "協助", "指導", "教", "說明", "解說"],
    "親切耐心": ["親切", "耐心", "貼心", "熱心", "細心", "溫柔"],
    "新手友善": ["第一次", "新手", "初學者", "不會", "看不懂", "零基礎"],
    "成品成功": ["成功", "完成", "拯救", "救了", "順利", "漂亮"],
    "環境氛圍": ["環境", "乾淨", "舒服", "氛圍", "明亮"],
    "回訪意願": ["下次", "再來", "還會來", "回訪"],
    "生日慶祝": ["生日", "壽星"],
}
ISSUE_THEME_KEYWORDS = {
    "服務態度": ["態度", "不耐煩", "兇", "跩", "口氣", "罵"],
    "清潔衛生": ["沒洗手", "洗乾淨", "油油", "清潔", "衛生", "髒"],
    "設備流程": ["設備", "機器", "烤箱", "平板", "充電", "流程", "結帳"],
    "空間動線": ["擁擠", "空間", "座位", "動線", "太小", "隔板"],
    "規則公告": ["公告", "標明", "規定", "12歲以下", "入場費"],
    "成品口味": ["乾", "硬", "像發糕", "不好吃", "失敗", "壓到"],
}
STORE_LABEL_PATTERN = re.compile(r"\(([^()]+)\)\s*$")
PUNCTUATION_PATTERN = re.compile(r"[，。！？、,.!?:;；\-~～\s]+")


def load_google_reviews_dashboard(source_dir: Path | None = None) -> dict:
    directory = (source_dir or DEFAULT_SOURCE_DIR).expanduser()
    files = sorted(directory.glob("*.xlsx"))
    raw_records: list[dict] = []
    files_summary: list[dict] = []
    errors: list[dict] = []

    for path in files:
        try:
            parsed = _parse_workbook(path)
        except Exception as exc:  # pragma: no cover - keep dashboard usable on partial failures
            errors.append({"file_name": path.name, "error": str(exc)})
            continue

        raw_records.extend(parsed["records"])
        files_summary.append(parsed["file_summary"])

    deduped_records, duplicate_count = _dedupe_records(raw_records)
    files_with_stats = _enrich_file_summaries(files_summary, raw_records, deduped_records)
    summary = _build_summary(raw_records, deduped_records, files_with_stats)
    stores = _build_store_summaries(deduped_records)
    recent_low_reviews = _build_recent_low_reviews(deduped_records)

    latest_source_update = max((item["modified_at"] for item in files_with_stats), default=None)

    return {
        "ok": True,
        "source_dir": str(directory),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_updated_at": latest_source_update,
        "file_count": len(files_with_stats),
        "raw_row_count": len(raw_records),
        "deduped_row_count": len(deduped_records),
        "duplicate_count": duplicate_count,
        "summary": summary,
        "stores": stores,
        "files": sorted(files_with_stats, key=lambda item: item["file_name"]),
        "recent_low_reviews": recent_low_reviews,
        "records": deduped_records,
        "errors": errors,
    }


def _parse_workbook(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    sheet_row_count = 0

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        row_iterator = worksheet.iter_rows(values_only=True)
        first_row = next(row_iterator, None)
        if first_row is None:
            continue

        column_map, first_data_row_number = _resolve_column_map(first_row)
        pending_rows: list[tuple[int, tuple]] = []
        if first_data_row_number == 1:
            pending_rows.append((1, first_row))

        for row_number, row in pending_rows:
            record = _build_record(row, column_map, path, sheet_name, row_number)
            if record:
                records.append(record)
                sheet_row_count += 1

        start_row_number = 2 if first_data_row_number == 2 else 2
        for index, row in enumerate(row_iterator, start=start_row_number):
            record = _build_record(row, column_map, path, sheet_name, index)
            if record:
                records.append(record)
                sheet_row_count += 1

    return {
        "records": records,
        "file_summary": {
            "file_name": path.name,
            "sheet_count": len(workbook.sheetnames),
            "raw_row_count": sheet_row_count,
            "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
        },
    }


def _resolve_column_map(first_row: tuple) -> tuple[dict[str, int], int]:
    normalized = [_normalize_header(cell) for cell in first_row]
    first_row_keys = {item for item in normalized if item}

    if HEADER_KEYS.issubset(first_row_keys):
        return {value: normalized.index(_normalize_header(value)) for value in TEXT_HEADER_ORDER}, 2

    if len(first_row) >= 5 and isinstance(first_row[1], datetime):
        return {key: index for index, key in enumerate(DEFAULT_HEADERLESS_ORDER)}, 1

    if len(first_row) >= 5 and _looks_like_datetime_text(first_row[3]):
        return {key: index for index, key in enumerate(TEXT_HEADER_ORDER)}, 1

    return {key: index for index, key in enumerate(DEFAULT_HEADERLESS_ORDER)}, 1


def _build_record(row: tuple, column_map: dict[str, int], path: Path, sheet_name: str, row_number: int) -> dict | None:
    values = {
        key: row[index] if index < len(row) else None
        for key, index in column_map.items()
    }

    title = _normalize_text(values.get("title"))
    published_at = _parse_datetime(values.get("publishedAtDate"))
    stars = _parse_stars(values.get("stars"))
    name = _normalize_text(values.get("name"))
    text = _normalize_text(values.get("text"))

    if not title or published_at is None or stars is None:
        return None

    store_label = _derive_store_label(title)
    positive_themes = _match_themes(text, POSITIVE_THEME_KEYWORDS)
    issue_themes = _match_themes(text, ISSUE_THEME_KEYWORDS)

    return {
        "title": title,
        "store_label": store_label,
        "stars": stars,
        "text": text,
        "name": name,
        "published_at": published_at.isoformat(timespec="seconds"),
        "month": published_at.strftime("%Y-%m"),
        "date": published_at.strftime("%Y-%m-%d"),
        "has_text": bool(text),
        "text_length": len(text),
        "is_low_star": stars <= 3,
        "is_five_star": stars == 5,
        "positive_themes": positive_themes,
        "issue_themes": issue_themes,
        "primary_issue_theme": issue_themes[0] if issue_themes else "",
        "file_name": path.name,
        "sheet_name": sheet_name,
        "row_number": row_number,
    }


def _dedupe_records(records: list[dict]) -> tuple[list[dict], int]:
    seen: set[str] = set()
    deduped: list[dict] = []

    for record in sorted(records, key=lambda item: (item["published_at"], item["store_label"], item["row_number"])):
        dedupe_key = "|".join(
            [
                record["title"].strip(),
                record["name"].strip(),
                record["published_at"],
                _normalize_for_dedup(record["text"]),
            ]
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        deduped.append(record)

    deduped.sort(key=lambda item: (item["published_at"], item["store_label"], item["row_number"]), reverse=True)
    return deduped, len(records) - len(deduped)


def _enrich_file_summaries(file_summaries: list[dict], raw_records: list[dict], deduped_records: list[dict]) -> list[dict]:
    raw_by_file = Counter(item["file_name"] for item in raw_records)
    deduped_by_file = Counter(item["file_name"] for item in deduped_records)

    enriched = []
    for item in file_summaries:
        file_name = item["file_name"]
        raw_count = raw_by_file.get(file_name, 0)
        deduped_count = deduped_by_file.get(file_name, 0)
        enriched.append(
            {
                **item,
                "raw_row_count": raw_count,
                "deduped_row_count": deduped_count,
                "duplicate_count": raw_count - deduped_count,
            }
        )
    return enriched


def _build_summary(raw_records: list[dict], deduped_records: list[dict], files_summary: list[dict]) -> dict:
    months = _aggregate_months(deduped_records)
    stores = _build_store_summaries(deduped_records)
    star_distribution = _aggregate_stars(deduped_records)
    positive_themes = _aggregate_theme_counts(deduped_records, "positive_themes")
    issue_themes = _aggregate_theme_counts(deduped_records, "issue_themes")
    low_reviews = [item for item in deduped_records if item["is_low_star"]]

    return {
        "store_count": len(stores),
        "raw_row_count": len(raw_records),
        "deduped_row_count": len(deduped_records),
        "duplicate_count": len(raw_records) - len(deduped_records),
        "duplicate_rate": _ratio(len(raw_records) - len(deduped_records), len(raw_records)),
        "average_star": round(mean(item["stars"] for item in deduped_records), 3) if deduped_records else 0,
        "text_rate": _ratio(sum(1 for item in deduped_records if item["has_text"]), len(deduped_records)),
        "five_star_rate": _ratio(sum(1 for item in deduped_records if item["is_five_star"]), len(deduped_records)),
        "low_star_count": len(low_reviews),
        "low_star_rate": _ratio(len(low_reviews), len(deduped_records)),
        "latest_review_at": max((item["published_at"] for item in deduped_records), default=None),
        "months": months,
        "stores": stores[:12],
        "star_distribution": star_distribution,
        "positive_themes": positive_themes,
        "issue_themes": issue_themes,
        "files": files_summary,
    }


def _build_store_summaries(records: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for record in records:
        grouped[record["store_label"]].append(record)

    max_reviews = max((len(items) for items in grouped.values()), default=1)
    summaries = []
    for store_label, items in grouped.items():
        review_count = len(items)
        low_star_count = sum(1 for item in items if item["is_low_star"])
        five_star_count = sum(1 for item in items if item["is_five_star"])
        text_count = sum(1 for item in items if item["has_text"])
        score = _health_score(
            average_star=mean(item["stars"] for item in items),
            low_star_rate=_ratio(low_star_count, review_count),
            text_rate=_ratio(text_count, review_count),
            review_count=review_count,
            max_reviews=max_reviews,
        )
        summaries.append(
            {
                "store_label": store_label,
                "review_count": review_count,
                "average_star": round(mean(item["stars"] for item in items), 3),
                "low_star_count": low_star_count,
                "low_star_rate": _ratio(low_star_count, review_count),
                "five_star_rate": _ratio(five_star_count, review_count),
                "text_rate": _ratio(text_count, review_count),
                "latest_review_at": max(item["published_at"] for item in items),
                "health_score": score,
                "top_positive_themes": _aggregate_theme_counts(items, "positive_themes")[:3],
                "top_issue_themes": _aggregate_theme_counts(items, "issue_themes")[:3],
            }
        )

    return sorted(summaries, key=lambda item: (-item["health_score"], -item["review_count"], item["store_label"]))


def _build_recent_low_reviews(records: list[dict]) -> list[dict]:
    low_reviews = [item for item in records if item["is_low_star"]]
    return sorted(low_reviews, key=lambda item: item["published_at"], reverse=True)[:40]


def _aggregate_months(records: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for record in records:
        grouped[record["month"]].append(record)

    months = []
    for month, items in grouped.items():
        months.append(
            {
                "month": month,
                "review_count": len(items),
                "average_star": round(mean(item["stars"] for item in items), 3),
                "low_star_count": sum(1 for item in items if item["is_low_star"]),
            }
        )
    return sorted(months, key=lambda item: item["month"])


def _aggregate_stars(records: list[dict]) -> list[dict]:
    counter = Counter(int(item["stars"]) for item in records)
    return [{"star": star, "review_count": counter.get(star, 0)} for star in range(1, 6)]


def _aggregate_theme_counts(records: list[dict], field_name: str) -> list[dict]:
    counter: Counter[str] = Counter()
    for record in records:
        counter.update(record[field_name])

    return [
        {"theme": theme, "review_count": count}
        for theme, count in counter.most_common()
    ]


def _health_score(average_star: float, low_star_rate: float, text_rate: float, review_count: int, max_reviews: int) -> int:
    star_score = min(max((average_star / 5) * 55, 0), 55)
    low_star_score = min(max((1 - low_star_rate) * 20, 0), 20)
    text_score = min(max(text_rate * 10, 0), 10)
    volume_score = min(max((review_count / max_reviews) * 15, 0), 15)
    return round(star_score + low_star_score + text_score + volume_score)


def _match_themes(text: str, theme_map: dict[str, list[str]]) -> list[str]:
    matches = []
    for theme, keywords in theme_map.items():
        if any(keyword in text for keyword in keywords):
            matches.append(theme)
    return matches


def _derive_store_label(title: str) -> str:
    matched = STORE_LABEL_PATTERN.search(title)
    if matched:
        return matched.group(1).strip()

    cleaned = title
    for phrase in ("【烘焙DIY】", "自己做 烘焙聚樂部", "自己做烘焙聚樂部", "吳寶春自己做"):
        cleaned = cleaned.replace(phrase, "")
    cleaned = cleaned.strip(" -")
    return cleaned or title


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\xa0", " ").strip()
    return str(value).replace("\xa0", " ").strip()


def _normalize_header(value: object) -> str:
    return _normalize_text(value).replace(" ", "").lower()


def _parse_datetime(value: object) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = _normalize_text(value)
    if not text:
        return None

    text = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _parse_stars(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        stars = int(float(value))
    except (TypeError, ValueError):
        return None
    return stars if 1 <= stars <= 5 else None


def _ratio(part: int | float, whole: int | float) -> float:
    if not whole:
        return 0.0
    return round(part / whole, 4)


def _looks_like_datetime_text(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    return "T" in text and any(char.isdigit() for char in text)


def _normalize_for_dedup(text: str) -> str:
    return PUNCTUATION_PATTERN.sub("", text).strip().lower()
